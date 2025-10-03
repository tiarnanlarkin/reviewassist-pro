from flask import Blueprint, request, jsonify, send_file
from src.models.user import db
from src.models.review import Review
from src.models.analytics import (
    AdvancedAnalytics, ReportTemplate, GeneratedReport, 
    PerformanceBenchmark, PredictiveInsight, AnalyticsEngine,
    ReportType, MetricType
)
from src.models.auth import AuthUser
from src.routes.auth import token_required
from datetime import datetime, timedelta, date
from sqlalchemy import func, and_, or_, desc
import json
import os
import tempfile
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.graphics.shapes import Drawing
from reportlab.graphics.charts.linecharts import HorizontalLineChart
from reportlab.graphics.charts.piecharts import Pie
import openpyxl
from openpyxl.chart import LineChart, PieChart, Reference
from openpyxl.styles import Font, Alignment, PatternFill
import csv

analytics_bp = Blueprint('analytics', __name__)

@analytics_bp.route('/dashboard-data', methods=['GET'])
@token_required
def get_dashboard_data(current_user):
    """Get comprehensive dashboard analytics data"""
    try:
        # Get date range parameters
        days = request.args.get('days', 30, type=int)
        end_date = datetime.utcnow().date()
        start_date = end_date - timedelta(days=days)
        
        # Basic metrics
        total_reviews = Review.query.count()
        period_reviews = Review.query.filter(
            Review.created_at >= start_date
        ).count()
        
        # Average rating
        avg_rating = db.session.query(func.avg(Review.rating)).scalar() or 0
        
        # Response rate
        responded_reviews = Review.query.filter(
            Review.ai_response.isnot(None)
        ).count() if hasattr(Review, 'ai_response') else 0
        response_rate = (responded_reviews / total_reviews * 100) if total_reviews > 0 else 0
        
        # Sentiment distribution
        sentiment_data = db.session.query(
            Review.sentiment,
            func.count(Review.id)
        ).group_by(Review.sentiment).all()
        
        sentiment_distribution = {
            'Positive': 0,
            'Neutral': 0,
            'Negative': 0
        }
        for sentiment, count in sentiment_data:
            if sentiment in sentiment_distribution:
                sentiment_distribution[sentiment] = count
        
        # Platform performance
        platform_data = db.session.query(
            Review.platform,
            func.count(Review.id),
            func.avg(Review.rating)
        ).group_by(Review.platform).all()
        
        platform_performance = []
        for platform, count, avg_rating in platform_data:
            platform_performance.append({
                'platform': platform,
                'review_count': count,
                'average_rating': round(float(avg_rating or 0), 1)
            })
        
        # Trend data (last 30 days)
        trend_data = []
        for i in range(30):
            day = end_date - timedelta(days=29-i)
            day_reviews = Review.query.filter(
                func.date(Review.created_at) == day
            ).count()
            trend_data.append({
                'date': day.isoformat(),
                'reviews': day_reviews
            })
        
        # Performance benchmarks
        benchmarks = PerformanceBenchmark.query.filter(
            PerformanceBenchmark.benchmark_date >= start_date
        ).all()
        
        # Predictive insights
        insights = PredictiveInsight.query.filter(
            PredictiveInsight.is_active == True
        ).order_by(desc(PredictiveInsight.confidence_score)).limit(5).all()
        
        return jsonify({
            'success': True,
            'data': {
                'overview': {
                    'total_reviews': total_reviews,
                    'period_reviews': period_reviews,
                    'average_rating': round(avg_rating, 1),
                    'response_rate': round(response_rate, 1)
                },
                'sentiment_distribution': sentiment_distribution,
                'platform_performance': platform_performance,
                'trend_data': trend_data,
                'benchmarks': [b.to_dict() for b in benchmarks],
                'insights': [i.to_dict() for i in insights]
            }
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@analytics_bp.route('/advanced-metrics', methods=['GET'])
@token_required
def get_advanced_metrics(current_user):
    """Get advanced analytics metrics with trend analysis"""
    try:
        # Get parameters
        metric_type = request.args.get('metric_type', 'all')
        days = request.args.get('days', 90, type=int)
        platform = request.args.get('platform')
        
        end_date = datetime.utcnow().date()
        start_date = end_date - timedelta(days=days)
        
        # Build query
        query = AdvancedAnalytics.query.filter(
            AdvancedAnalytics.date >= start_date,
            AdvancedAnalytics.date <= end_date
        )
        
        if metric_type != 'all':
            query = query.filter(AdvancedAnalytics.metric_type == metric_type)
        
        if platform:
            query = query.filter(AdvancedAnalytics.platform == platform)
        
        metrics = query.order_by(AdvancedAnalytics.date).all()
        
        # Group by metric type and analyze trends
        grouped_metrics = {}
        for metric in metrics:
            key = metric.metric_type.value
            if key not in grouped_metrics:
                grouped_metrics[key] = []
            grouped_metrics[key].append({
                'date': metric.date.isoformat(),
                'value': metric.value,
                'platform': metric.platform,
                'additional_data': metric.additional_data
            })
        
        # Calculate trends for each metric
        trend_analysis = {}
        for metric_key, data_points in grouped_metrics.items():
            values = [point['value'] for point in data_points]
            trend = AnalyticsEngine.calculate_trend(values)
            anomalies = AnalyticsEngine.detect_anomalies(values)
            forecast = AnalyticsEngine.forecast_next_period(values, 7)  # 7 days ahead
            
            trend_analysis[metric_key] = {
                'data': data_points,
                'trend': trend,
                'anomalies': anomalies,
                'forecast': forecast,
                'statistics': {
                    'mean': sum(values) / len(values) if values else 0,
                    'min': min(values) if values else 0,
                    'max': max(values) if values else 0,
                    'latest': values[-1] if values else 0
                }
            }
        
        return jsonify({
            'success': True,
            'data': {
                'metrics': trend_analysis,
                'date_range': {
                    'start': start_date.isoformat(),
                    'end': end_date.isoformat()
                }
            }
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@analytics_bp.route('/generate-insights', methods=['POST'])
@token_required
def generate_insights(current_user):
    """Generate predictive insights based on current data"""
    try:
        # Get recent review data for analysis
        days = 90
        end_date = datetime.utcnow().date()
        start_date = end_date - timedelta(days=days)
        
        # Get daily review counts
        daily_reviews = db.session.query(
            func.date(Review.created_at).label('date'),
            func.count(Review.id).label('count'),
            func.avg(Review.rating).label('avg_rating')
        ).filter(
            Review.created_at >= start_date
        ).group_by(func.date(Review.created_at)).all()
        
        insights = []
        
        if daily_reviews:
            # Analyze review volume trend
            review_counts = [r.count for r in daily_reviews]
            volume_trend = AnalyticsEngine.calculate_trend(review_counts)
            
            if volume_trend['direction'] == 'increasing' and volume_trend['strength'] > 0.5:
                insights.append({
                    'type': 'trend',
                    'title': 'Increasing Review Volume',
                    'description': f"Review volume has increased by {volume_trend['change_percent']:.1f}% over the last {days} days. This indicates growing customer engagement.",
                    'confidence': 0.8,
                    'recommendations': [
                        'Ensure adequate staff to handle increased response volume',
                        'Monitor response times to maintain quality',
                        'Consider implementing automated acknowledgments'
                    ]
                })
            
            # Analyze rating trends
            ratings = [float(r.avg_rating or 0) for r in daily_reviews]
            rating_trend = AnalyticsEngine.calculate_trend(ratings)
            
            if rating_trend['direction'] == 'decreasing' and rating_trend['strength'] > 0.3:
                insights.append({
                    'type': 'alert',
                    'title': 'Declining Rating Trend',
                    'description': f"Average ratings have decreased by {abs(rating_trend['change_percent']):.1f}% recently. Immediate attention may be needed.",
                    'confidence': 0.9,
                    'recommendations': [
                        'Review recent negative feedback for common issues',
                        'Implement service improvement measures',
                        'Increase proactive customer outreach'
                    ]
                })
            
            # Forecast next week's review volume
            forecast = AnalyticsEngine.forecast_next_period(review_counts, 7)
            if forecast:
                insights.append({
                    'type': 'forecast',
                    'title': 'Weekly Review Volume Forecast',
                    'description': f"Based on current trends, expect approximately {int(forecast)} reviews in the next 7 days.",
                    'confidence': 0.7,
                    'predicted_value': forecast,
                    'recommendations': [
                        'Prepare response templates for expected volume',
                        'Schedule staff accordingly',
                        'Monitor actual vs. predicted performance'
                    ]
                })
        
        # Save insights to database
        for insight_data in insights:
            insight = PredictiveInsight(
                insight_type=insight_data['type'],
                title=insight_data['title'],
                description=insight_data['description'],
                confidence_score=insight_data['confidence'],
                predicted_value=insight_data.get('predicted_value'),
                predicted_date=end_date + timedelta(days=7) if insight_data['type'] == 'forecast' else None,
                action_recommendations=insight_data['recommendations']
            )
            db.session.add(insight)
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'data': {
                'insights_generated': len(insights),
                'insights': insights
            }
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@analytics_bp.route('/generate-report', methods=['POST'])
@token_required
def generate_report(current_user):
    """Generate comprehensive PDF or Excel report"""
    try:
        data = request.get_json()
        report_type = data.get('report_type', 'monthly')
        file_format = data.get('format', 'pdf')
        date_range = data.get('date_range', {})
        filters = data.get('filters', {})
        
        # Parse date range
        if date_range:
            start_date = datetime.strptime(date_range['start'], '%Y-%m-%d').date()
            end_date = datetime.strptime(date_range['end'], '%Y-%m-%d').date()
        else:
            end_date = datetime.utcnow().date()
            if report_type == 'weekly':
                start_date = end_date - timedelta(days=7)
            elif report_type == 'monthly':
                start_date = end_date - timedelta(days=30)
            elif report_type == 'quarterly':
                start_date = end_date - timedelta(days=90)
            else:
                start_date = end_date - timedelta(days=365)
        
        # Gather report data
        report_data = gather_report_data(start_date, end_date, filters)
        
        # Generate file
        if file_format == 'pdf':
            file_path = generate_pdf_report(report_data, start_date, end_date, report_type)
        elif file_format == 'xlsx':
            file_path = generate_excel_report(report_data, start_date, end_date, report_type)
        else:
            file_path = generate_csv_report(report_data, start_date, end_date, report_type)
        
        # Save report record
        report = GeneratedReport(
            title=f"{report_type.title()} Report - {start_date} to {end_date}",
            report_type=ReportType(report_type),
            date_range_start=start_date,
            date_range_end=end_date,
            filters_applied=filters,
            data_snapshot=report_data,
            file_path=file_path,
            file_type=file_format,
            generated_by=current_user.id,
            expires_at=datetime.utcnow() + timedelta(days=30)  # Auto-cleanup after 30 days
        )
        db.session.add(report)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'data': {
                'report_id': report.id,
                'file_path': file_path,
                'download_url': f'/api/analytics/download-report/{report.id}'
            }
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@analytics_bp.route('/download-report/<int:report_id>', methods=['GET'])
@token_required
def download_report(current_user, report_id):
    """Download generated report"""
    try:
        report = GeneratedReport.query.get_or_404(report_id)
        
        # Update download count
        report.download_count += 1
        db.session.commit()
        
        return send_file(
            report.file_path,
            as_attachment=True,
            download_name=f"{report.title}.{report.file_type}"
        )
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

def gather_report_data(start_date, end_date, filters):
    """Gather comprehensive data for report generation"""
    # Base query
    query = Review.query.filter(
        Review.created_at >= start_date,
        Review.created_at <= end_date
    )
    
    # Apply filters
    if filters.get('platform'):
        query = query.filter(Review.platform == filters['platform'])
    if filters.get('rating_min'):
        query = query.filter(Review.rating >= filters['rating_min'])
    if filters.get('rating_max'):
        query = query.filter(Review.rating <= filters['rating_max'])
    if filters.get('sentiment'):
        query = query.filter(Review.sentiment == filters['sentiment'])
    
    reviews = query.all()
    
    # Calculate metrics
    total_reviews = len(reviews)
    avg_rating = sum(r.rating for r in reviews) / total_reviews if total_reviews > 0 else 0
    
    # Sentiment distribution
    sentiment_counts = {'Positive': 0, 'Neutral': 0, 'Negative': 0}
    for review in reviews:
        if review.sentiment in sentiment_counts:
            sentiment_counts[review.sentiment] += 1
    
    # Platform breakdown
    platform_data = {}
    for review in reviews:
        if review.platform not in platform_data:
            platform_data[review.platform] = {'count': 0, 'ratings': []}
        platform_data[review.platform]['count'] += 1
        platform_data[review.platform]['ratings'].append(review.rating)
    
    # Calculate platform averages
    for platform in platform_data:
        ratings = platform_data[platform]['ratings']
        platform_data[platform]['avg_rating'] = sum(ratings) / len(ratings) if ratings else 0
    
    # Daily trends
    daily_data = {}
    for review in reviews:
        day = review.created_at.date()
        if day not in daily_data:
            daily_data[day] = {'count': 0, 'ratings': []}
        daily_data[day]['count'] += 1
        daily_data[day]['ratings'].append(review.rating)
    
    return {
        'summary': {
            'total_reviews': total_reviews,
            'average_rating': round(avg_rating, 2),
            'date_range': {
                'start': start_date.isoformat(),
                'end': end_date.isoformat()
            }
        },
        'sentiment_distribution': sentiment_counts,
        'platform_performance': platform_data,
        'daily_trends': {
            str(day): {
                'count': data['count'],
                'avg_rating': sum(data['ratings']) / len(data['ratings']) if data['ratings'] else 0
            }
            for day, data in daily_data.items()
        },
        'reviews': [
            {
                'id': r.id,
                'platform': r.platform,
                'rating': r.rating,
                'sentiment': r.sentiment,
                'created_at': r.created_at.isoformat(),
                'reviewer_name': r.reviewer_name,
                'content': r.content[:200] + '...' if len(r.content) > 200 else r.content
            }
            for r in reviews[:100]  # Limit to first 100 for report
        ]
    }

def generate_pdf_report(data, start_date, end_date, report_type):
    """Generate PDF report"""
    # Create temporary file
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.pdf')
    temp_file.close()
    
    # Create PDF document
    doc = SimpleDocTemplate(temp_file.name, pagesize=A4)
    styles = getSampleStyleSheet()
    story = []
    
    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        spaceAfter=30,
        alignment=1  # Center alignment
    )
    story.append(Paragraph(f"ReviewAssist Pro - {report_type.title()} Report", title_style))
    story.append(Paragraph(f"Period: {start_date} to {end_date}", styles['Normal']))
    story.append(Spacer(1, 20))
    
    # Summary section
    story.append(Paragraph("Executive Summary", styles['Heading2']))
    summary_data = [
        ['Metric', 'Value'],
        ['Total Reviews', str(data['summary']['total_reviews'])],
        ['Average Rating', f"{data['summary']['average_rating']}/5.0"],
        ['Report Period', f"{start_date} to {end_date}"]
    ]
    
    summary_table = Table(summary_data)
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 14),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    story.append(summary_table)
    story.append(Spacer(1, 20))
    
    # Platform performance
    story.append(Paragraph("Platform Performance", styles['Heading2']))
    platform_data = [['Platform', 'Reviews', 'Avg Rating']]
    for platform, metrics in data['platform_performance'].items():
        platform_data.append([
            platform,
            str(metrics['count']),
            f"{metrics['avg_rating']:.1f}"
        ])
    
    platform_table = Table(platform_data)
    platform_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    story.append(platform_table)
    
    # Build PDF
    doc.build(story)
    
    return temp_file.name

def generate_excel_report(data, start_date, end_date, report_type):
    """Generate Excel report"""
    # Create temporary file
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    temp_file.close()
    
    # Create workbook
    wb = openpyxl.Workbook()
    
    # Summary sheet
    ws_summary = wb.active
    ws_summary.title = "Summary"
    
    # Add title
    ws_summary['A1'] = f"ReviewAssist Pro - {report_type.title()} Report"
    ws_summary['A1'].font = Font(size=16, bold=True)
    ws_summary['A2'] = f"Period: {start_date} to {end_date}"
    
    # Add summary data
    ws_summary['A4'] = "Metric"
    ws_summary['B4'] = "Value"
    ws_summary['A5'] = "Total Reviews"
    ws_summary['B5'] = data['summary']['total_reviews']
    ws_summary['A6'] = "Average Rating"
    ws_summary['B6'] = data['summary']['average_rating']
    
    # Platform performance sheet
    ws_platforms = wb.create_sheet("Platform Performance")
    ws_platforms['A1'] = "Platform"
    ws_platforms['B1'] = "Review Count"
    ws_platforms['C1'] = "Average Rating"
    
    row = 2
    for platform, metrics in data['platform_performance'].items():
        ws_platforms[f'A{row}'] = platform
        ws_platforms[f'B{row}'] = metrics['count']
        ws_platforms[f'C{row}'] = metrics['avg_rating']
        row += 1
    
    # Daily trends sheet
    ws_trends = wb.create_sheet("Daily Trends")
    ws_trends['A1'] = "Date"
    ws_trends['B1'] = "Review Count"
    ws_trends['C1'] = "Average Rating"
    
    row = 2
    for date_str, metrics in data['daily_trends'].items():
        ws_trends[f'A{row}'] = date_str
        ws_trends[f'B{row}'] = metrics['count']
        ws_trends[f'C{row}'] = metrics['avg_rating']
        row += 1
    
    # Save workbook
    wb.save(temp_file.name)
    
    return temp_file.name

def generate_csv_report(data, start_date, end_date, report_type):
    """Generate CSV report"""
    # Create temporary file
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.csv', mode='w', newline='')
    
    writer = csv.writer(temp_file)
    
    # Write header
    writer.writerow([f"ReviewAssist Pro - {report_type.title()} Report"])
    writer.writerow([f"Period: {start_date} to {end_date}"])
    writer.writerow([])
    
    # Write summary
    writer.writerow(["Summary"])
    writer.writerow(["Total Reviews", data['summary']['total_reviews']])
    writer.writerow(["Average Rating", data['summary']['average_rating']])
    writer.writerow([])
    
    # Write platform performance
    writer.writerow(["Platform Performance"])
    writer.writerow(["Platform", "Review Count", "Average Rating"])
    for platform, metrics in data['platform_performance'].items():
        writer.writerow([platform, metrics['count'], metrics['avg_rating']])
    
    temp_file.close()
    return temp_file.name

